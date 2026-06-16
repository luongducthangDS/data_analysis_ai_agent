"""
Export endpoints — work with any domain dashboard (not e-commerce specific).

GET /api/dashboard/{session_id}/export.xlsx       → Excel workbook (KPIs + Top 10)
GET /api/dashboard/{session_id}/export-chart/{chart_id}.png → Chart PNG (requires kaleido)
"""
from __future__ import annotations

import io
import json
import logging

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from backend.app.api.deps import get_current_user, get_session
from backend.app.services.storage import DatasetSession

_log = logging.getLogger(__name__)
router = APIRouter()


def _build_xlsx(session: DatasetSession) -> bytes:
    """Build an Excel workbook from the cached AI dashboard response."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cached = getattr(session, "_dashboard_cache", None)
    if cached is None:
        raise ValueError("Dashboard not computed yet. Call GET /api/dashboard first.")

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    ALERT_FILL  = PatternFill("solid", fgColor="FFCCCC")

    def _hrow(ws, cols: list[str]) -> None:
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def _autowidth(ws) -> None:
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    # ── Sheet 1: KPI Summary ──────────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "KPI"
    domain = cached.platform or "Dashboard"
    _hrow(ws_kpi, ["Chỉ số", "Giá trị", "So sánh", "Công thức"])
    for i, card in enumerate(cached.kpi_cards, start=2):
        ws_kpi.append([card.label, card.value, card.delta or "—", card.formula or ""])
        if card.is_alert:
            for cell in ws_kpi[i]:
                cell.fill = ALERT_FILL
    _autowidth(ws_kpi)

    # ── Sheet 2: Top 10 breakdown ─────────────────────────────────────────────
    ws_top = wb.create_sheet("Top 10")
    _hrow(ws_top, ["Hạng", "Nhóm", "Giá trị"])
    for row in cached.top_products:
        ws_top.append([
            row.get("rank", ""),
            str(row.get("name", "")),
            row.get("value", ""),
        ])
    _autowidth(ws_top)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────

@router.get("/api/dashboard/{session_id}/export.xlsx")
def export_dashboard_xlsx(
    session_id: str,
    _session: DatasetSession = Depends(get_session),
    _user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Download AI dashboard data as an Excel workbook."""
    if getattr(_session, "_dashboard_cache", None) is None:
        raise HTTPException(
            status_code=422,
            detail="Dashboard not computed yet. Open the Dashboard tab first.",
        )
    try:
        xlsx_bytes = _build_xlsx(_session)
    except Exception as exc:
        _log.exception("export_xlsx: failed session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="dashboard_{session_id[:8]}.xlsx"'},
    )


@router.get("/api/dashboard/{session_id}/export-chart/{chart_id}.png")
def export_chart_png(
    session_id: str,
    chart_id: str,
    _session: DatasetSession = Depends(get_session),
    _user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Download a single dashboard chart as a PNG image (requires kaleido)."""
    charts_raw: list[dict] = getattr(_session, "_dashboard_charts_raw", None) or []
    chart_data = next((c for c in charts_raw if c.get("chart_id") == chart_id), None)
    if not chart_data:
        raise HTTPException(
            status_code=404,
            detail=f"Chart '{chart_id}' not found. Open Dashboard tab first.",
        )

    try:
        import plotly.graph_objects as go
        import plotly.io as pio

        fig = go.Figure(json.loads(json.dumps(chart_data["plotly_json"])))
        png_bytes = pio.to_image(fig, format="png", width=900, height=420)
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="PNG export requires kaleido. Add 'kaleido>=0.2.1' to requirements.",
        )
    except Exception as exc:
        _log.exception("export_chart_png: failed chart=%s session=%s", chart_id, session_id)
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return StreamingResponse(
        io.BytesIO(png_bytes),
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="chart_{chart_id[:8]}.png"'},
    )
